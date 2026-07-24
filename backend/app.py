import os
import csv
from datetime import datetime, timedelta
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, abort, send_file
from io import StringIO, BytesIO
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URI', 'sqlite:///' + os.path.join(os.path.dirname(os.path.abspath(__file__)), 'equipment.db'))
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ---------- 数据模型 ----------
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    name = db.Column(db.String(50))
    role = db.Column(db.String(20), default='普通员工')
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=True)

class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    manager = db.Column(db.String(50))
    start_date = db.Column(db.Date)
    estimated_end_date = db.Column(db.Date)
    remark = db.Column(db.Text)
    is_virtual = db.Column(db.Boolean, default=False)

class Device(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    category = db.Column(db.String(50))
    brand = db.Column(db.String(50))
    model = db.Column(db.String(100))
    serial_number = db.Column(db.String(100), unique=True)
    asset_number = db.Column(db.String(100), unique=True)
    ownership = db.Column(db.String(10), default='自有')
    status = db.Column(db.String(20), default='闲置')
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'))
    purchase_date = db.Column(db.Date)
    purchase_price = db.Column(db.Float)
    supplier = db.Column(db.String(100))
    warranty_end = db.Column(db.Date)
    lease_id = db.Column(db.Integer, db.ForeignKey('lease_contract.id'))
    remark = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    project = db.relationship('Project', backref='devices')
    lease = db.relationship('LeaseContract', backref='devices')

class DeviceCategory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class LeaseContract(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    supplier = db.Column(db.String(100))
    contract_number = db.Column(db.String(100))
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    monthly_cost = db.Column(db.Float)
    total_cost = db.Column(db.Float)
    status = db.Column(db.String(20), default='正常')
    remark = db.Column(db.Text)

class OperationLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    username = db.Column(db.String(50))
    action = db.Column(db.String(50))
    device_serial = db.Column(db.String(100))
    detail = db.Column(db.Text)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

class PurchaseOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    purchase_date = db.Column(db.Date)
    supplier = db.Column(db.String(100))
    total_amount = db.Column(db.Float)
    remark = db.Column(db.Text)

# ---------- 初始化数据库 ----------
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

with app.app_context():
    db.create_all()
    
    try:
        exists = db.session.execute(db.select(db.exists().where(Project.id == 1))).scalar()
    except Exception:
        exists = False
    if not exists:
        virtual = Project(id=1, name='公司待用库', is_virtual=True, manager='系统')
        db.session.add(virtual)
        db.session.commit()
    
    try:
        admin_exists = db.session.execute(db.select(db.exists().where(User.username == 'admin'))).scalar()
    except Exception:
        admin_exists = False
    if not admin_exists:
        admin = User(username='admin', name='系统管理员', role='超级管理员',
                     password_hash=generate_password_hash('admin123'))
        db.session.add(admin)
        db.session.commit()
    
    default_categories = ['台式主机', '笔记本', '显示器', '打印机', '扫描仪']
    for cat_name in default_categories:
        try:
            exists = db.session.execute(db.select(db.exists().where(DeviceCategory.name == cat_name))).scalar()
        except Exception:
            exists = False
        if not exists:
            cat = DeviceCategory(name=cat_name)
            db.session.add(cat)
            db.session.commit()

# ---------- 权限装饰器 ----------
def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if current_user.role not in roles:
                flash('没有访问权限', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ---------- 路由 ----------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('用户名或密码错误', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    total_devices = Device.query.count()
    in_use = Device.query.filter_by(status='在用').count()
    idle = Device.query.filter_by(status='闲置').count()
    lease_expiring_soon = Device.query.filter(
        Device.ownership == '租赁',
        Device.lease.has(LeaseContract.end_date <= datetime.now().date() + timedelta(days=30))
    ).count() if Device.query.filter(Device.ownership=='租赁').first() else 0  # 简单处理避免无租赁时报错
    logs = OperationLog.query.order_by(OperationLog.timestamp.desc()).limit(10).all()
    return render_template('dashboard.html',
                           total=total_devices, in_use=in_use, idle=idle,
                           expiring=lease_expiring_soon, logs=logs)

@app.route('/device')
@login_required
def device_list():
    projects = Project.query.all()
    categories = DeviceCategory.query.order_by(DeviceCategory.name).all()
    return render_template('device_list.html', projects=projects, categories=categories)

@app.route('/device/api')
@login_required
def device_api():
    if current_user.role == '项目经理' and current_user.project_id:
        devices_query = Device.query.filter_by(project_id=current_user.project_id)
    else:
        devices_query = Device.query
    search = request.args.get('search')
    project_id = request.args.get('project_id', type=int)
    status = request.args.get('status')
    category = request.args.get('category')
    ownership = request.args.get('ownership')
    if search:
        devices_query = devices_query.filter(
            (Device.serial_number.contains(search)) |
            (Device.asset_number.contains(search)) |
            (Device.brand.contains(search)) |
            (Device.model.contains(search))
        )
    if project_id:
        devices_query = devices_query.filter_by(project_id=project_id)
    if status:
        devices_query = devices_query.filter_by(status=status)
    if category:
        devices_query = devices_query.filter_by(category=category)
    if ownership:
        devices_query = devices_query.filter_by(ownership=ownership)
    devices = devices_query.order_by(Device.created_at.desc()).all()
    
    result = []
    for d in devices:
        result.append({
            'id': d.id,
            'category': d.category,
            'brand': d.brand,
            'model': d.model,
            'serial_number': d.serial_number,
            'asset_number': d.asset_number,
            'project_name': d.project.name if d.project else '-',
            'status': d.status,
            'ownership': d.ownership,
            'is_admin': current_user.role == '超级管理员'
        })
    return jsonify(result)

@app.route('/device/add', methods=['POST'])
@login_required
def device_add():
    serial = request.form.get('serial_number')
    if Device.query.filter_by(serial_number=serial).first():
        flash('序列号已存在', 'danger')
        return redirect(url_for('device_list'))
    device = Device(
        category=request.form.get('category'),
        brand=request.form.get('brand'),
        model=request.form.get('model'),
        serial_number=serial,
        asset_number=request.form.get('asset_number'),
        ownership=request.form.get('ownership', '自有'),
        status='闲置',
        project_id=request.form.get('project_id') or 1,
        purchase_date=datetime.strptime(request.form.get('purchase_date'), '%Y-%m-%d') if request.form.get('purchase_date') else None,
        purchase_price=float(request.form['purchase_price']) if request.form.get('purchase_price') else None,
        supplier=request.form.get('supplier'),
        warranty_end=datetime.strptime(request.form.get('warranty_end'), '%Y-%m-%d') if request.form.get('warranty_end') else None,
        remark=request.form.get('remark')
    )
    db.session.add(device)
    log = OperationLog(user_id=current_user.id, username=current_user.name,
                       action='创建', device_serial=device.serial_number,
                       detail=f'新增设备 {device.category} {device.brand} {device.model}')
    db.session.add(log)
    db.session.commit()
    flash('设备添加成功', 'success')
    referrer = request.headers.get('Referer', '')
    if 'project/walk' in referrer:
        return redirect(referrer)
    return redirect(url_for('device_list'))

@app.route('/device/<int:id>/edit', methods=['POST'])
@login_required
def device_edit(id):
    device = Device.query.get_or_404(id)
    if current_user.role == '项目经理' and device.project_id != current_user.project_id:
        abort(403)
    device.category = request.form['category']
    device.brand = request.form['brand']
    device.model = request.form['model']
    device.serial_number = request.form['serial_number']
    device.asset_number = request.form['asset_number']
    device.ownership = request.form['ownership']
    device.purchase_date = datetime.strptime(request.form['purchase_date'], '%Y-%m-%d') if request.form['purchase_date'] else None
    device.purchase_price = float(request.form['purchase_price']) if request.form.get('purchase_price') else None
    device.supplier = request.form['supplier']
    device.warranty_end = datetime.strptime(request.form['warranty_end'], '%Y-%m-%d') if request.form.get('warranty_end') else None
    device.remark = request.form['remark']
    log = OperationLog(user_id=current_user.id, username=current_user.name,
                       action='编辑', device_serial=device.serial_number,
                       detail='更新设备信息')
    db.session.add(log)
    db.session.commit()
    flash('设备信息更新成功', 'success')
    return redirect(url_for('device_detail', id=id))

@app.route('/device/<int:id>')
@login_required
def device_detail(id):
    device = Device.query.get_or_404(id)
    logs = OperationLog.query.filter_by(device_serial=device.serial_number).order_by(OperationLog.timestamp.desc()).all()
    categories = DeviceCategory.query.order_by(DeviceCategory.name).all()
    return render_template('device_detail.html', device=device, logs=logs, Project=Project, categories=categories)

@app.route('/device/<int:id>/status', methods=['POST'])
@login_required
def device_status_change(id):
    device = Device.query.get_or_404(id)
    action = request.form['action']
    if action == '调拨':
        target_project_id = request.form.get('target_project_id')
        receiver = request.form.get('receiver')
        device.status = '在用'
        device.project_id = int(target_project_id)
        detail = f'调拨至 {Project.query.get(target_project_id).name}，接收人 {receiver}'
    elif action == '送修':
        device.status = '维修中'
        detail = '设备送修'
    elif action == '归还':
        device.status = '闲置'
        device.project_id = 1
        detail = '归还至公司待用库'
    elif action == '报废':
        device.status = '已报废'
        detail = '设备报废'
    db.session.add(device)
    log = OperationLog(user_id=current_user.id, username=current_user.name,
                       action=action, device_serial=device.serial_number,
                       detail=detail)
    db.session.add(log)
    db.session.commit()
    flash('操作成功', 'success')
    return redirect(url_for('device_detail', id=id))

@app.route('/device/<int:id>/delete', methods=['POST'])
@login_required
@role_required('超级管理员')
def device_delete(id):
    device = Device.query.get_or_404(id)
    db.session.delete(device)
    log = OperationLog(user_id=current_user.id, username=current_user.name,
                       action='删除', device_serial=device.serial_number,
                       detail=f'删除设备 {device.category} {device.brand} {device.model}')
    db.session.add(log)
    db.session.commit()
    flash('设备已删除', 'success')
    return redirect(url_for('device_list'))

@app.route('/device/<int:id>/delete_api', methods=['DELETE'])
@login_required
@role_required('超级管理员')
def device_delete_api(id):
    device = Device.query.get_or_404(id)
    db.session.delete(device)
    log = OperationLog(user_id=current_user.id, username=current_user.name,
                       action='删除', device_serial=device.serial_number,
                       detail=f'删除设备 {device.category} {device.brand} {device.model}')
    db.session.add(log)
    db.session.commit()
    return jsonify({'success': True, 'message': '设备已删除'})

@app.route('/device/import/template')
@login_required
def device_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = '设备导入模板'
    
    headers = ['类别', '品牌', '型号', '序列号', '资产编号', '属性', '采购日期', '采购单价', '供应商', '质保截止', '备注']
    ws.append(headers)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    ws.append(['台式主机', '联想', 'ThinkPad', 'SN-001', 'AST-001', '自有', '2024-01-15', 5000.00, '联想代理商', '2027-01-15', '备注信息'])
    ws.append(['笔记本', '华为', 'MateBook', 'SN-002', 'AST-002', '自有', '2024-02-20', 6500.00, '华为代理商', '2027-02-20', ''])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='设备导入模板.xlsx',
        as_attachment=True
    )

@app.route('/device/import', methods=['POST'])
@login_required
def device_import():
    if 'file' not in request.files:
        flash('请选择文件', 'danger')
        return redirect(url_for('device_list'))
    
    file = request.files['file']
    if file.filename == '':
        flash('请选择文件', 'danger')
        return redirect(url_for('device_list'))
    
    if not file.filename.endswith('.xlsx'):
        flash('请上传xlsx格式的Excel文件', 'danger')
        return redirect(url_for('device_list'))
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        required_fields = ['类别', '品牌', '型号', '序列号']
        for field in required_fields:
            if field not in headers:
                flash(f'Excel文件缺少必需字段: {field}', 'danger')
                return redirect(url_for('device_list'))
        
        field_indices = {header: headers.index(header) for header in headers}
        
        success_count = 0
        fail_count = 0
        errors = []
        project_id = request.form.get('project_id', 1)
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                serial = str(row[field_indices['序列号']]).strip() if row[field_indices['序列号']] else ''
                if not serial:
                    fail_count += 1
                    errors.append(f'第{row_num}行: 序列号不能为空')
                    continue
                
                if Device.query.filter_by(serial_number=serial).first():
                    fail_count += 1
                    errors.append(f'第{row_num}行: 序列号 {serial} 已存在')
                    continue
                
                category = str(row[field_indices['类别']]).strip() if row[field_indices['类别']] else ''
                brand = str(row[field_indices['品牌']]).strip() if row[field_indices['品牌']] else ''
                model = str(row[field_indices['型号']]).strip() if row[field_indices['型号']] else ''
                
                if not category or not brand or not model:
                    fail_count += 1
                    errors.append(f'第{row_num}行: 必填字段不完整')
                    continue
                
                asset_number = str(row[field_indices['资产编号']]).strip() if row[field_indices['资产编号']] else None
                ownership = str(row[field_indices['属性']]).strip() if row[field_indices['属性']] else '自有'
                
                purchase_date_val = row[field_indices['采购日期']]
                purchase_date = None
                if purchase_date_val:
                    if isinstance(purchase_date_val, datetime):
                        purchase_date = purchase_date_val.date()
                    else:
                        try:
                            purchase_date = datetime.strptime(str(purchase_date_val), '%Y-%m-%d').date()
                        except:
                            purchase_date = None
                
                purchase_price_val = row[field_indices['采购单价']]
                purchase_price = float(purchase_price_val) if purchase_price_val else None
                
                supplier = str(row[field_indices['供应商']]).strip() if row[field_indices['供应商']] else None
                
                warranty_end_val = row[field_indices['质保截止']]
                warranty_end = None
                if warranty_end_val:
                    if isinstance(warranty_end_val, datetime):
                        warranty_end = warranty_end_val.date()
                    else:
                        try:
                            warranty_end = datetime.strptime(str(warranty_end_val), '%Y-%m-%d').date()
                        except:
                            warranty_end = None
                
                remark = str(row[field_indices['备注']]).strip() if row[field_indices['备注']] else None
                
                device = Device(
                    category=category,
                    brand=brand,
                    model=model,
                    serial_number=serial,
                    asset_number=asset_number,
                    ownership=ownership,
                    status='闲置',
                    project_id=int(project_id),
                    purchase_date=purchase_date,
                    purchase_price=purchase_price,
                    supplier=supplier,
                    warranty_end=warranty_end,
                    remark=remark
                )
                db.session.add(device)
                
                log = OperationLog(user_id=current_user.id, username=current_user.name,
                                   action='创建', device_serial=device.serial_number,
                                   detail=f'批量导入设备 {device.category} {device.brand} {device.model}')
                db.session.add(log)
                
                success_count += 1
            except Exception as e:
                fail_count += 1
                errors.append(f'第{row_num}行: {str(e)}')
        
        db.session.commit()
        
        message = f'批量导入完成，成功 {success_count} 台'
        if fail_count > 0:
            message += f'，失败 {fail_count} 台'
            for error in errors[:5]:
                message += f'; {error}'
            if len(errors) > 5:
                message += f'; ...还有 {len(errors) - 5} 条错误'
        
        flash(message, 'success' if success_count > 0 else 'danger')
        
    except Exception as e:
        flash(f'导入失败: {str(e)}', 'danger')
    
    referrer = request.headers.get('Referer', '')
    if 'project/walk' in referrer:
        return redirect(referrer)
    return redirect(url_for('device_list'))

@app.route('/device/batch', methods=['POST'])
@login_required
def device_batch_action():
    action = request.form['action']
    device_ids_str = request.form.get('device_ids', '')
    device_ids = [id.strip() for id in device_ids_str.split(',') if id.strip()]
    
    if not device_ids:
        flash('请选择要操作的设备', 'danger')
        return redirect(url_for('device_list'))
    
    success_count = 0
    fail_count = 0
    
    for device_id in device_ids:
        try:
            device = Device.query.get(int(device_id))
            if not device:
                fail_count += 1
                continue
            
            if action == '调拨':
                if device.status == '已报废':
                    fail_count += 1
                    continue
                target_project_id = request.form.get('target_project_id')
                receiver = request.form.get('receiver')
                device.status = '在用'
                device.project_id = int(target_project_id)
                detail = f'批量调拨至 {Project.query.get(target_project_id).name}，接收人 {receiver}'
                
            elif action == '送修':
                if device.status in ['维修中', '已报废']:
                    fail_count += 1
                    continue
                device.status = '维修中'
                detail = '批量送修'
                
            elif action == '归还':
                if device.status != '在用':
                    fail_count += 1
                    continue
                device.status = '闲置'
                device.project_id = 1
                detail = '批量归还至公司待用库'
                
            elif action == '报废':
                if device.status == '已报废':
                    fail_count += 1
                    continue
                device.status = '已报废'
                detail = '批量报废'
            
            else:
                fail_count += 1
                continue
            
            db.session.add(device)
            log = OperationLog(user_id=current_user.id, username=current_user.name,
                               action=action, device_serial=device.serial_number,
                               detail=detail)
            db.session.add(log)
            success_count += 1
            
        except Exception as e:
            fail_count += 1
    
    db.session.commit()
    
    message = f'批量操作完成，成功 {success_count} 台'
    if fail_count > 0:
        message += f'，失败 {fail_count} 台'
    flash(message, 'success')
    return redirect(url_for('device_list'))

@app.route('/project')
@login_required
def project_list():
    projects = Project.query.filter(Project.is_virtual == False).all()
    return render_template('project_list.html', projects=projects)

@app.route('/project/add', methods=['POST'])
@login_required
@role_required('超级管理员')
def project_add():
    name = request.form['name']
    if Project.query.filter_by(name=name).first():
        flash('项目名称已存在', 'danger')
        return redirect(url_for('project_list'))
    start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date() if request.form.get('start_date') else None
    estimated_end_date = datetime.strptime(request.form['estimated_end_date'], '%Y-%m-%d').date() if request.form.get('estimated_end_date') else None
    project = Project(
        name=name, 
        manager=request.form['manager'], 
        start_date=start_date,
        estimated_end_date=estimated_end_date,
        remark=request.form['remark']
    )
    db.session.add(project)
    db.session.commit()
    flash('项目创建成功', 'success')
    return redirect(url_for('project_list'))

@app.route('/project/<int:id>/edit', methods=['POST'])
@login_required
@role_required('超级管理员')
def project_edit(id):
    project = Project.query.get_or_404(id)
    if project.is_virtual:
        abort(403)
    project.name = request.form['name']
    project.manager = request.form['manager']
    project.start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date() if request.form.get('start_date') else None
    project.estimated_end_date = datetime.strptime(request.form['estimated_end_date'], '%Y-%m-%d').date() if request.form.get('estimated_end_date') else None
    project.remark = request.form['remark']
    db.session.commit()
    flash('项目更新成功', 'success')
    return redirect(url_for('project_list'))

@app.route('/project/<int:id>/delete', methods=['POST'])
@login_required
@role_required('超级管理员')
def project_delete(id):
    project = Project.query.get_or_404(id)
    if project.is_virtual or Device.query.filter_by(project_id=id).count() > 0:
        flash('不能删除有设备的项目', 'danger')
    else:
        db.session.delete(project)
        db.session.commit()
        flash('项目已删除', 'success')
    return redirect(url_for('project_list'))

@app.route('/project/<int:id>/export')
@login_required
def project_export(id):
    project = Project.query.get_or_404(id)
    if current_user.role == '项目经理' and project.id != current_user.project_id:
        abort(403)
    
    devices = Device.query.filter_by(project_id=id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = project.name + '设备清单'
    
    headers = ['类别', '品牌', '型号', '序列号', '资产编号', '状态', '属性', 
               '采购日期', '采购单价', '供应商', '质保截止', '备注']
    ws.append(headers)
    
    for d in devices:
        ws.append([
            d.category or '',
            d.brand or '',
            d.model or '',
            d.serial_number or '',
            d.asset_number or '',
            d.status or '',
            d.ownership or '',
            d.purchase_date.strftime('%Y-%m-%d') if d.purchase_date else '',
            d.purchase_price or '',
            d.supplier or '',
            d.warranty_end.strftime('%Y-%m-%d') if d.warranty_end else '',
            d.remark or ''
        ])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{project.name}_设备清单_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename,
        as_attachment=True
    )

@app.route('/project/walk')
@login_required
def project_walk():
    if current_user.role == '项目经理':
        projects = Project.query.filter_by(id=current_user.project_id).all()
    else:
        projects = Project.query.all()
    project_id = request.args.get('project_id', type=int)
    if not project_id and projects:
        project_id = projects[0].id
    project = Project.query.get_or_404(project_id)
    if current_user.role == '项目经理' and project.id != current_user.project_id:
        abort(403)
    devices = Device.query.filter_by(project_id=project.id).all()
    categories = DeviceCategory.query.order_by(DeviceCategory.name).all()
    return render_template('project_walk.html', projects=projects, current_project=project, devices=devices, categories=categories)

@app.route('/purchase')
@login_required
@role_required('超级管理员', '财务/库管')
def purchase():
    orders = PurchaseOrder.query.order_by(PurchaseOrder.purchase_date.desc()).all()
    categories = DeviceCategory.query.order_by(DeviceCategory.name).all()
    return render_template('purchase.html', orders=orders, categories=categories)

@app.route('/purchase/<int:id>')
@login_required
@role_required('超级管理员', '财务/库管')
def purchase_detail(id):
    order = PurchaseOrder.query.get_or_404(id)
    devices = Device.query.filter(
        Device.purchase_date == order.purchase_date,
        Device.supplier == order.supplier
    ).order_by(Device.id).all()
    return render_template('purchase_detail.html', order=order, devices=devices)

@app.route('/purchase/add', methods=['POST'])
@login_required
@role_required('超级管理员', '财务/库管')
def purchase_add():
    purchase_date = datetime.strptime(request.form['purchase_date'], '%Y-%m-%d')
    supplier = request.form['supplier']
    remark = request.form['remark']
    categories = request.form.getlist('category')
    brands = request.form.getlist('brand')
    models = request.form.getlist('model')
    quantities = request.form.getlist('quantity')
    prices = request.form.getlist('price')
    order = PurchaseOrder(purchase_date=purchase_date, supplier=supplier, total_amount=0, remark=remark)
    db.session.add(order)
    total_amount = 0
    for i in range(len(categories)):
        qty = int(quantities[i])
        price = float(prices[i])
        total_amount += qty * price
        for _ in range(qty):
            device = Device(
                category=categories[i],
                brand=brands[i],
                model=models[i],
                serial_number=f'NEW-{datetime.now().strftime("%Y%m%d%H%M%S")}-{i}',
                asset_number='',
                ownership='自有',
                status='闲置',
                project_id=1,
                purchase_date=purchase_date,
                purchase_price=price,
                supplier=supplier
            )
            db.session.add(device)
    order.total_amount = total_amount
    log = OperationLog(user_id=current_user.id, username=current_user.name,
                       action='采购入库', device_serial='批量',
                       detail=f'采购单 {order.id}，供应商 {supplier}，总金额 {total_amount}')
    db.session.add(log)
    db.session.commit()
    flash('采购入库完成，请及时补充设备序列号', 'success')
    return redirect(url_for('purchase'))

@app.route('/lease')
@login_required
def lease():
    lease_devices = Device.query.filter_by(ownership='租赁').all()
    projects = Project.query.all()
    categories = DeviceCategory.query.order_by(DeviceCategory.name).all()
    now = datetime.now()
    return render_template('lease.html', lease_devices=lease_devices, projects=projects, categories=categories, now=now)

@app.route('/lease/<int:id>')
@login_required
def lease_detail(id):
    contract = LeaseContract.query.get_or_404(id)
    return jsonify({
        'id': contract.id,
        'supplier': contract.supplier,
        'contract_number': contract.contract_number,
        'start_date': contract.start_date.strftime('%Y-%m-%d') if contract.start_date else '',
        'end_date': contract.end_date.strftime('%Y-%m-%d') if contract.end_date else '',
        'monthly_cost': contract.monthly_cost,
        'total_cost': contract.total_cost,
        'remark': contract.remark
    })

@app.route('/lease/add', methods=['POST'])
@login_required
def lease_add():
    contract = LeaseContract(
        supplier=request.form['supplier'],
        contract_number=request.form['contract_number'],
        start_date=datetime.strptime(request.form['start_date'], '%Y-%m-%d') if request.form['start_date'] else None,
        end_date=datetime.strptime(request.form['end_date'], '%Y-%m-%d') if request.form['end_date'] else None,
        monthly_cost=float(request.form['monthly_cost']) if request.form.get('monthly_cost') else None,
        total_cost=float(request.form['total_cost']) if request.form.get('total_cost') else None,
        remark=request.form['remark']
    )
    db.session.add(contract)
    db.session.commit()
    
    device = Device(
        category=request.form['category'],
        brand=request.form['brand'],
        model=request.form['model'],
        serial_number=request.form['serial_number'],
        asset_number=request.form.get('asset_number'),
        project_id=int(request.form['project_id']) if request.form.get('project_id') else None,
        purchase_date=None,
        purchase_price=None,
        supplier=request.form['supplier'],
        warranty_end=None,
        status='正常使用',
        ownership='租赁',
        lease_id=contract.id,
        remark=request.form.get('remark')
    )
    db.session.add(device)
    db.session.commit()
    
    flash('租赁设备已登记', 'success')
    return redirect(url_for('lease'))

@app.route('/lease/<int:id>/edit', methods=['POST'])
@login_required
def lease_edit(id):
    contract = LeaseContract.query.get_or_404(id)
    contract.supplier = request.form['supplier']
    contract.contract_number = request.form['contract_number']
    contract.start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d') if request.form['start_date'] else None
    contract.end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d') if request.form['end_date'] else None
    contract.monthly_cost = float(request.form['monthly_cost']) if request.form.get('monthly_cost') else None
    contract.total_cost = float(request.form['total_cost']) if request.form.get('total_cost') else None
    contract.remark = request.form['remark']
    db.session.commit()
    flash('租赁合同已更新', 'success')
    return redirect(url_for('lease'))

@app.route('/lease/<int:id>/delete', methods=['POST'])
@login_required
def lease_delete(id):
    contract = LeaseContract.query.get_or_404(id)
    if Device.query.filter_by(lease_id=id).count() > 0:
        flash('无法删除关联设备的合同', 'danger')
    else:
        db.session.delete(contract)
        db.session.commit()
        flash('租赁合同已删除', 'success')
    return redirect(url_for('lease'))

@app.route('/report')
@login_required
def report():
    logs = OperationLog.query.order_by(OperationLog.timestamp.desc()).limit(20).all()
    projects = Project.query.filter(Project.is_virtual == False).all()
    return render_template('report.html', logs=logs, projects=projects)

@app.route('/report/export')
@login_required
def report_export():
    logs = OperationLog.query.order_by(OperationLog.timestamp.desc()).all()
    
    output = StringIO()
    output.write('\ufeff')
    
    writer = csv.writer(output)
    writer.writerow(['时间', '操作人', '操作类型', '设备序列号', '详情'])
    
    for log in logs:
        writer.writerow([
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.username,
            log.action,
            log.device_serial or '',
            log.detail or ''
        ])
    
    output.seek(0)
    output_bytes = output.getvalue().encode('utf-8')
    output = BytesIO(output_bytes)
    
    return send_file(
        output,
        mimetype='text/csv',
        download_name=f'操作日志_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv',
        as_attachment=True
    )

@app.route('/report/project_assets/<int:project_id>')
@login_required
def report_project_assets(project_id):
    project = Project.query.get_or_404(project_id)
    devices = Device.query.filter_by(project_id=project_id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = project.name + '资产明细表'
    
    headers = ['类别', '品牌', '型号', '序列号', '资产编号', '状态', '属性', 
               '采购日期', '采购单价', '供应商', '质保截止', '备注']
    ws.append(headers)
    
    for d in devices:
        ws.append([
            d.category or '',
            d.brand or '',
            d.model or '',
            d.serial_number or '',
            d.asset_number or '',
            d.status or '',
            d.ownership or '',
            d.purchase_date.strftime('%Y-%m-%d') if d.purchase_date else '',
            d.purchase_price or '',
            d.supplier or '',
            d.warranty_end.strftime('%Y-%m-%d') if d.warranty_end else '',
            d.remark or ''
        ])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{project.name}_资产明细表_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename,
        as_attachment=True
    )

@app.route('/report/fixed_assets')
@login_required
def report_fixed_assets():
    devices = Device.query.filter_by(ownership='自有').all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = '固定资产总账'
    
    headers = ['所属项目', '类别', '品牌', '型号', '序列号', '资产编号', 
               '采购日期', '采购单价', '供应商', '质保截止', '状态']
    ws.append(headers)
    
    total_value = 0
    for d in devices:
        ws.append([
            d.project.name if d.project else '-',
            d.category or '',
            d.brand or '',
            d.model or '',
            d.serial_number or '',
            d.asset_number or '',
            d.purchase_date.strftime('%Y-%m-%d') if d.purchase_date else '',
            d.purchase_price or '',
            d.supplier or '',
            d.warranty_end.strftime('%Y-%m-%d') if d.warranty_end else '',
            d.status or ''
        ])
        if d.purchase_price:
            total_value += d.purchase_price
    
    ws.append([])
    ws.append(['总计资产价值', f'¥{total_value:.2f}'])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"固定资产总账_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename,
        as_attachment=True
    )

@app.route('/report/lease_cost')
@login_required
def report_lease_cost():
    devices = Device.query.filter_by(ownership='租赁').all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = '租赁费用汇总'
    
    headers = ['所属项目', '类别', '品牌', '型号', '序列号', '供应商',
               '合同号', '开始日期', '到期日期', '月租金', '预估总租金']
    ws.append(headers)
    
    total_monthly = 0
    for d in devices:
        contract = d.lease
        monthly_cost = contract.monthly_cost if contract else 0
        total_monthly += monthly_cost
        
        ws.append([
            d.project.name if d.project else '-',
            d.category or '',
            d.brand or '',
            d.model or '',
            d.serial_number or '',
            contract.supplier if contract else '',
            contract.contract_number if contract else '',
            contract.start_date.strftime('%Y-%m-%d') if contract and contract.start_date else '',
            contract.end_date.strftime('%Y-%m-%d') if contract and contract.end_date else '',
            monthly_cost or '',
            contract.total_cost if contract else ''
        ])
    
    ws.append([])
    ws.append(['月租金总计', f'¥{total_monthly:.2f}'])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"租赁费用汇总_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename,
        as_attachment=True
    )

@app.route('/system/users')
@login_required
@role_required('超级管理员')
def system_users():
    users = User.query.all()
    projects = Project.query.filter_by(is_virtual=False).all()
    return render_template('system_users.html', users=users, projects=projects)

@app.route('/system/user/reset_password/<int:id>', methods=['POST'])
@login_required
@role_required('超级管理员')
def user_reset_password(id):
    user = User.query.get_or_404(id)
    user.password_hash = generate_password_hash('123456')
    db.session.commit()
    flash(f'{user.name} 的密码已重置为 123456', 'success')
    return redirect(url_for('system_users'))

@app.route('/system/user/toggle_status/<int:id>', methods=['POST'])
@login_required
@role_required('超级管理员')
def user_toggle_status(id):
    user = User.query.get_or_404(id)
    if user.username == 'admin':
        flash('不能禁用管理员账号', 'danger')
    else:
        user.is_active = not getattr(user, 'is_active', True)
        db.session.commit()
        status = '启用' if user.is_active else '禁用'
        flash(f'{user.name} 已{status}', 'success')
    return redirect(url_for('system_users'))

@app.route('/system/user/add', methods=['POST'])
@login_required
@role_required('超级管理员')
def user_add():
    username = request.form['username']
    if User.query.filter_by(username=username).first():
        flash('用户名已存在', 'danger')
        return redirect(url_for('system_users'))
    user = User(
        username=username,
        name=request.form['name'],
        role=request.form['role'],
        project_id=request.form.get('project_id') if request.form['role'] == '项目经理' else None,
        password_hash=generate_password_hash(request.form['password'])
    )
    db.session.add(user)
    db.session.commit()
    flash('用户创建成功', 'success')
    return redirect(url_for('system_users'))

@app.route('/system/config')
@login_required
@role_required('超级管理员')
def system_config():
    categories = DeviceCategory.query.order_by(DeviceCategory.name).all()
    return render_template('system_config.html', categories=categories)

@app.route('/system/config/category/add', methods=['POST'])
@login_required
@role_required('超级管理员')
def add_category():
    name = request.form['name'].strip()
    if name and not DeviceCategory.query.filter_by(name=name).first():
        category = DeviceCategory(name=name)
        db.session.add(category)
        db.session.commit()
        flash(f'类别 "{name}" 已添加', 'success')
    else:
        flash('类别名称无效或已存在', 'danger')
    return redirect(url_for('system_config'))

@app.route('/system/config/category/delete/<string:name>', methods=['POST'])
@login_required
@role_required('超级管理员')
def delete_category(name):
    category = DeviceCategory.query.filter_by(name=name).first()
    if not category:
        flash('类别不存在', 'danger')
    elif Device.query.filter_by(category=name).count() > 0:
        flash(f'无法删除已有设备使用的类别 "{name}"', 'danger')
    else:
        db.session.delete(category)
        db.session.commit()
        flash(f'类别 "{name}" 已删除', 'success')
    return redirect(url_for('system_config'))

@app.route('/api/categories')
@login_required
def categories_api():
    categories = DeviceCategory.query.order_by(DeviceCategory.name).all()
    return jsonify([{'id': c.id, 'name': c.name} for c in categories])

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(debug=debug_mode, port=int(os.environ.get('PORT', 5000)))